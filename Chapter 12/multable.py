#!python3
#this program creates a multiplication table, please create a bat file to run this
#usage: multable.bat num

import openpyxl, sys
from openpyxl.styles import Font

num = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active
for rows in range(1, num+1):
    sheet.cell(row = rows+1, column=1).value=rows
    sheet.cell(row = rows+1, column=1).font = Font(bold=True)
    for i in range(1, num+1):
        sheet.cell(row=rows+1, column=i+1).value = rows*i
        sheet.cell(row=1, column=i+1).value = i
        sheet.cell(row=1, column=i+1).font = Font(bold=True)

wb.save('C:\\Users\\glend\\OneDrive\\Desktop\\python stuff\\chapter 12\\Multable%s.xlsx' %num)
print('done making table.')
